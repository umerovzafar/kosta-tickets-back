

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from infrastructure.models import AbsenceDay, ScheduleEmployee


def _find_header_row(values_rows: list[tuple]) -> int:
    for i, row in enumerate(values_rows):
        if not row or len(row) < 3:
            continue
        c0 = row[0]
        c1 = row[1]
        if c0 is not None and str(c0).strip() in {"№", "N", "No"}:
            if c1 is not None and "ФИО" in str(c1).upper().replace(" ", ""):
                return i
        if str(c0).strip() == "№" and c1 and "ФИО" in str(c1):
            return i
    raise ValueError("Не найдена строка заголовка с «№» и «ФИО»")


def _date_columns(header_row: tuple) -> list[tuple[int, date]]:
    out: list[tuple[int, date]] = []
    for idx, val in enumerate(header_row):
        if isinstance(val, datetime):
            out.append((idx, val.date()))
        elif isinstance(val, date):
            out.append((idx, val))
    return out


def _row_values(ws, row_1based: int, max_col: int) -> tuple:
    return tuple(ws.cell(row=row_1based, column=c).value for c in range(1, max_col + 1))


def _validate_year_matches_columns(year: int, col_dates: list[tuple[int, date]]) -> None:
    for _, d in col_dates:
        if d.year != year:
            raise ValueError(
                f"Год в файле ({d.year}) не совпадает с выбранным годом импорта ({year}). "
                "Загрузите файл за нужный год или укажите другой year."
            )


def import_schedule_from_workbook(
    session: Session,
    *,
    year: int,
    sheet_name: str | None,
    source: Path | bytes,
) -> tuple[int, int]:

    if isinstance(source, Path):
        wb = load_workbook(source, data_only=True, read_only=False)
    else:
        wb = load_workbook(BytesIO(source), data_only=True, read_only=False)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        max_row = ws.max_row
        max_col = ws.max_column
        preview: list[tuple] = []
        for r in range(1, min(max_row, 30) + 1):
            preview.append(_row_values(ws, r, max_col))
        hi = _find_header_row(preview)
        header_1based = hi + 1
        header_row = preview[hi]
        col_dates = _date_columns(header_row)
        if not col_dates:
            raise ValueError("В строке заголовка нет колонок с датами")

        _validate_year_matches_columns(year, col_dates)

        data_start_1based = header_1based + 2

        session.execute(delete(ScheduleEmployee).where(ScheduleEmployee.year == year))

        emp_count = 0
        day_count = 0
        for r in range(data_start_1based, max_row + 1):
            row = _row_values(ws, r, max_col)
            name_cell = row[1] if len(row) > 1 else None
            if name_cell is None or not str(name_cell).strip():
                continue
            full_name = str(name_cell).strip()
            if full_name.upper() in {"ФИО", "ИТОГО", "ВСЕГО"}:
                continue

            excel_no = row[0]
            row_no: int | None = None
            if isinstance(excel_no, int):
                row_no = excel_no
            elif isinstance(excel_no, float) and excel_no == int(excel_no):
                row_no = int(excel_no)
            elif excel_no is not None and str(excel_no).strip().isdigit():
                row_no = int(str(excel_no).strip())

            note = row[2] if len(row) > 2 else None
            note_s = str(note).strip() if note is not None and str(note).strip() else None

            emp = ScheduleEmployee(
                year=year,
                excel_row_no=row_no,
                full_name=full_name,
                planned_period_note=note_s,
            )
            session.add(emp)
            session.flush()

            for col_idx, d in col_dates:
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                if raw is None:
                    continue
                code: int | None = None
                if isinstance(raw, int) and 1 <= raw <= 5:
                    code = raw
                elif isinstance(raw, float) and raw == int(raw) and 1 <= int(raw) <= 5:
                    code = int(raw)
                elif isinstance(raw, str):
                    m = re.match(r"^\s*([1-5])\s*$", raw)
                    if m:
                        code = int(m.group(1))
                if code is not None:
                    session.add(AbsenceDay(employee_id=emp.id, absence_on=d, kind_code=code))
                    day_count += 1
            emp_count += 1

        session.commit()
        return emp_count, day_count
    finally:
        wb.close()
