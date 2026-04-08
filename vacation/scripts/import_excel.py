"""
Импорт графика отсутствий из Excel (как «График_отпусков_работников_на_2026г.xlsx»).

Ожидается лист с заголовком строки: № | ФИО | (примечание) | даты по колонкам.
В ячейках дней — коды 1–5 (легенда: ежегодный отпуск, болезнь, day off, командировка, удалёнка).

Запуск (из каталога vacation/, с поднятой БД):
  set DATABASE_URL=postgresql://vacation:123456@localhost:5432/kosta_vacation
  python -m scripts.import_excel "C:\\path\\to\\file.xlsx" --year 2026

В Docker-сети хост для Postgres с Windows/Mac часто host.docker.internal.
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, delete
from sqlalchemy.orm import Session, sessionmaker

# пакет при запуске как python -m scripts.import_excel из каталога vacation
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from infrastructure.models import AbsenceDay, ScheduleEmployee  # noqa: E402


def _sync_engine_url(database_url: str) -> str:
    u = database_url.strip()
    if u.startswith("postgresql+asyncpg://"):
        return u.replace("postgresql+asyncpg://", "postgresql://", 1)
    return u


def _find_header_row(values_rows: list[tuple]) -> int:
    for i, row in enumerate(values_rows):
        if not row or len(row) < 3:
            continue
        c0 = row[0]
        c1 = row[1]
        if c0 is not None and str(c0).strip() in {"№", "N", "No"}:
            if c1 is not None and "ФИО" in str(c1).upper().replace(" ", ""):
                return i
        if str(c0).strip() == "№" and c1 and "ФИО" in str(c1):
            return i
    raise ValueError("Не найдена строка заголовка с «№» и «ФИО»")


def _date_columns(header_row: tuple) -> list[tuple[int, date]]:
    out: list[tuple[int, date]] = []
    for idx, val in enumerate(header_row):
        if isinstance(val, datetime):
            out.append((idx, val.date()))
        elif isinstance(val, date):
            out.append((idx, val))
    return out


def _row_values(ws, row_1based: int, max_col: int) -> tuple:
    return tuple(ws.cell(row=row_1based, column=c).value for c in range(1, max_col + 1))


def import_workbook(session: Session, xlsx_path: Path, year: int, sheet_name: str | None) -> tuple[int, int]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        max_row = ws.max_row
        max_col = ws.max_column
        preview: list[tuple] = []
        for r in range(1, min(max_row, 30) + 1):
            preview.append(_row_values(ws, r, max_col))
        hi = _find_header_row(preview)
        header_1based = hi + 1
        header_row = preview[hi]
        col_dates = _date_columns(header_row)
        if not col_dates:
            raise ValueError("В строке заголовка нет колонок с датами")

        # Строка дней недели сразу под датами; данные — со следующей
        data_start_1based = header_1based + 2

        session.execute(delete(ScheduleEmployee))

        emp_count = 0
        day_count = 0
        for r in range(data_start_1based, max_row + 1):
            row = _row_values(ws, r, max_col)
            name_cell = row[1] if len(row) > 1 else None
            if name_cell is None or not str(name_cell).strip():
                continue
            full_name = str(name_cell).strip()
            if full_name.upper() in {"ФИО", "ИТОГО", "ВСЕГО"}:
                continue

            excel_no = row[0]
            row_no: int | None = None
            if isinstance(excel_no, int):
                row_no = excel_no
            elif isinstance(excel_no, float) and excel_no == int(excel_no):
                row_no = int(excel_no)
            elif excel_no is not None and str(excel_no).strip().isdigit():
                row_no = int(str(excel_no).strip())

            note = row[2] if len(row) > 2 else None
            note_s = str(note).strip() if note is not None and str(note).strip() else None

            emp = ScheduleEmployee(
                year=year,
                excel_row_no=row_no,
                full_name=full_name,
                planned_period_note=note_s,
            )
            session.add(emp)
            session.flush()

            for col_idx, d in col_dates:
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                if raw is None:
                    continue
                code: int | None = None
                if isinstance(raw, int) and 1 <= raw <= 5:
                    code = raw
                elif isinstance(raw, float) and raw == int(raw) and 1 <= int(raw) <= 5:
                    code = int(raw)
                elif isinstance(raw, str):
                    m = re.match(r"^\s*([1-5])\s*$", raw)
                    if m:
                        code = int(m.group(1))
                if code is not None:
                    session.add(AbsenceDay(employee_id=emp.id, absence_on=d, kind_code=code))
                    day_count += 1
            emp_count += 1

        session.commit()
        return emp_count, day_count
    finally:
        wb.close()


def main() -> None:
    p = argparse.ArgumentParser(description="Импорт графика отсутствий из Excel в БД vacation.")
    p.add_argument("xlsx", type=Path, help="Путь к .xlsx")
    p.add_argument("--year", type=int, default=2026)
    p.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый)")
    args = p.parse_args()

    import os

    url = (os.environ.get("DATABASE_URL") or "").strip()
    if not url:
        print("Задайте DATABASE_URL (postgresql://...)", file=sys.stderr)
        sys.exit(1)
    if not args.xlsx.is_file():
        print(f"Файл не найден: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    engine = create_engine(_sync_engine_url(url), echo=False)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)
    with SessionLocal() as session:
        n_emp, n_day = import_workbook(session, args.xlsx, args.year, args.sheet)
    print(f"Импорт завершён: сотрудников={n_emp}, отмеченных дней={n_day}, год={args.year}")


if __name__ == "__main__":
    main()
