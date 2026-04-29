

from __future__ import annotations

import csv
import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

from starlette.responses import Response


def _filename(report_type: str, group_by: str | None, date_from: date, date_to: date, fmt: str) -> str:

    parts = [report_type]
    if group_by:
        parts.append(group_by)
    parts.append(date_from.isoformat())
    parts.append(date_to.isoformat())
    return "_".join(parts) + f".{fmt}"


def _collect_fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    keys: list[str] = []
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def _json_default(o: Any) -> Any:

    if hasattr(o, "model_dump"):
        try:
            return o.model_dump(mode="json")
        except TypeError:
            return o.model_dump()
    if hasattr(o, "dict"):
        return o.dict()
    if isinstance(o, (date, datetime)):
        return o.isoformat()
    if isinstance(o, Decimal):
        return float(o)
    return str(o)


def _cell_value_for_export(val: Any) -> Any:

    if val is None:
        return ""
    if val is True or val is False:
        return val
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, (date, datetime)):
        return val
    if isinstance(val, str):
        return val
    if isinstance(val, (dict, list, tuple, set)):
        try:
            return json.dumps(val, ensure_ascii=False, default=_json_default)
        except TypeError:
            return str(val)
    return str(val)


def _row_to_plain_dict(row: Any) -> dict[str, Any]:

    if isinstance(row, dict):
        return dict(row)
    model_dump = getattr(row, "model_dump", None)
    if callable(model_dump):
        try:
            return model_dump(mode="json")
        except TypeError:
            return model_dump()
    dict_m = getattr(row, "dict", None)
    if callable(dict_m):
        return dict_m()
    mapping = getattr(row, "_mapping", None)
    if mapping is not None:
        return dict(mapping)
    raise TypeError(f"Report row must be dict-like, got {type(row)!r}")


def _sanitize_export_rows(rows: list[Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        d = _row_to_plain_dict(row)
        out.append({str(k): _cell_value_for_export(v) for k, v in d.items()})
    return out


def export_csv(
    rows: list[Any],
    report_type: str,
    group_by: str | None,
    date_from: date,
    date_to: date,
) -> Response:
    buf = StringIO()
    safe_rows = _sanitize_export_rows(rows)
    if safe_rows:
        fieldnames = _collect_fieldnames(safe_rows)
        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", restval="")
        writer.writeheader()
        writer.writerows(safe_rows)

    content = ("\ufeff" + buf.getvalue()).encode("utf-8")
    fname = _filename(report_type, group_by, date_from, date_to, "csv")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def export_xlsx(
    rows: list[Any],
    report_type: str,
    group_by: str | None,
    date_from: date,
    date_to: date,
) -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    safe_rows = _sanitize_export_rows(rows)
    fieldnames = _collect_fieldnames(safe_rows) if safe_rows else []


    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_humanize_header(name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


    for row_idx, row in enumerate(safe_rows, start=2):
        for col_idx, name in enumerate(fieldnames, start=1):
            raw = row.get(name, "")
            if raw is None:
                raw = ""
            val = _cell_value_for_export(raw)
            ws.cell(row=row_idx, column=col_idx, value=val)


    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = _filename(report_type, group_by, date_from, date_to, "xlsx")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _humanize_header(name: str) -> str:

    if " " in name or "?" in name:
        return name
    return name.replace("_", " ").title()


def export_report(
    rows: list[Any],
    fmt: str,
    report_type: str,
    group_by: str | None,
    date_from: date,
    date_to: date,
) -> Response:

    if fmt == "xlsx":
        return export_xlsx(rows, report_type, group_by, date_from, date_to)
    return export_csv(rows, report_type, group_by, date_from, date_to)
